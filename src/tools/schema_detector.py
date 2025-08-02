"""Schema detection and inference for structured data using LLM analysis."""

import pandas as pd
import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from openai import OpenAI
from loguru import logger
import json
import asyncio

from src.config import settings
from src.tools.web_search import WebSearchTool


class SchemaDetector:
    """Intelligent schema detection for Excel files using LLM analysis."""

    def __init__(self):
        self.openai_client = OpenAI(api_key=settings.openai_api_key)
        self.web_search = WebSearchTool()

    async def detect_sheet_schemas(
        self,
        file_path: Path,
        context_pdfs: List[str] = None,
        web_source_url: str = None,
    ) -> Dict[str, Any]:
        """Detect schemas for all sheets in an Excel file."""

        try:
            # Load Excel file
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Get context information
            context_info = await self._gather_context(
                file_path, context_pdfs, web_source_url
            )

            schemas = {}

            for sheet_name in workbook.sheetnames:
                logger.info(f"Analyzing schema for sheet: {sheet_name}")

                # Extract sheet data
                sheet_data = self._extract_sheet_sample(workbook[sheet_name])

                # Detect multiple tables within sheet
                tables = self._detect_tables_in_sheet(workbook[sheet_name])

                # Analyze schema with LLM
                schema = await self._analyze_schema_with_llm(
                    sheet_name, sheet_data, tables, context_info
                )

                schemas[sheet_name] = schema

            return {
                "file_path": str(file_path),
                "schemas": schemas,
                "context": context_info,
                "total_sheets": len(workbook.sheetnames),
            }

        except Exception as e:
            logger.error(f"Schema detection failed for {file_path}: {e}")
            return {"error": str(e)}

    def _extract_sheet_sample(self, sheet, max_rows: int = 4) -> List[List[str]]:
        """Extract a minimal sample of data from the sheet (header + 3 data rows)."""
        data = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            row_data = [str(cell) if cell is not None else "" for cell in row]
            if any(cell.strip() for cell in row_data):  # Skip empty rows
                data.append(row_data)
        return data

    def _detect_tables_in_sheet(self, sheet) -> List[Dict[str, Any]]:
        """Detect multiple tables within a single sheet."""
        tables = []
        current_table = None

        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            row_data = [str(cell) if cell is not None else "" for cell in row]

            # Check if this looks like a header row
            if self._is_header_row(row_data):
                # Start new table
                if current_table:
                    tables.append(current_table)

                current_table = {
                    "start_row": i,
                    "headers": [cell for cell in row_data if cell.strip()],
                    "data_rows": [],
                }

            elif current_table and any(cell.strip() for cell in row_data):
                # Add data row to current table
                current_table["data_rows"].append(row_data)

            elif current_table and not any(cell.strip() for cell in row_data):
                # Empty row might end current table
                if len(current_table["data_rows"]) > 0:
                    current_table["end_row"] = i - 1

        # Add the last table
        if current_table and len(current_table["data_rows"]) > 0:
            tables.append(current_table)

        return tables

    def _is_header_row(self, row_data: List[str]) -> bool:
        """Heuristic to detect if a row is likely a header."""
        filled_cells = [cell for cell in row_data if cell.strip()]

        if len(filled_cells) < 2:
            return False

        # Check for typical header patterns
        has_text = any(any(c.isalpha() for c in cell) for cell in filled_cells)
        has_caps = any(cell.isupper() or cell.istitle() for cell in filled_cells)
        no_long_numbers = all(
            not (cell.replace(".", "").replace(",", "").isdigit() and len(cell) > 6)
            for cell in filled_cells
        )

        return has_text and (has_caps or no_long_numbers)

    async def _gather_context(
        self,
        file_path: Path,
        context_pdfs: List[str] = None,
        web_source_url: str = None,
    ) -> Dict[str, Any]:
        """Gather contextual information to understand schema."""

        context = {
            "file_name": file_path.name,
            "file_stem": file_path.stem,
            "pdf_context": [],
            "web_context": {},
            "directory_context": [],
        }

        # Search for related PDFs in the same directory
        if context_pdfs:
            # TODO: Extract relevant text from PDFs
            context["pdf_context"] = context_pdfs

        # Get web context if URL provided
        if web_source_url:
            try:
                web_results = await self.web_search.search(
                    f"site:{web_source_url} {file_path.stem}", max_results=3
                )
                context["web_context"] = {
                    "source_url": web_source_url,
                    "related_pages": web_results,
                }
            except Exception as e:
                logger.warning(f"Failed to get web context: {e}")

        # Get directory context
        try:
            parent_dir = file_path.parent
            related_files = [f.name for f in parent_dir.glob("*.pdf")][:5]
            context["directory_context"] = related_files
        except Exception as e:
            logger.warning(f"Failed to get directory context: {e}")

        return context

    async def _analyze_schema_with_llm(
        self,
        sheet_name: str,
        sheet_data: List[List[str]],
        tables: List[Dict[str, Any]],
        context: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Use LLM to analyze and infer schema structure."""

        prompt = self._build_schema_analysis_prompt(
            sheet_name, sheet_data, tables, context
        )

        try:
            response = self.openai_client.chat.completions.create(
                model=settings.llm_extraction_model,
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert data analyst specializing in Excel schema detection and analysis. Your task is to analyze Excel sheet data and infer meaningful schemas.",
                    },
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                response_format={"type": "json_object"},
            )

            schema_analysis = json.loads(response.choices[0].message.content)
            return schema_analysis

        except Exception as e:
            logger.error(f"LLM schema analysis failed: {e}")
            return {
                "error": str(e),
                "fallback_schema": self._create_fallback_schema(sheet_data, tables),
            }

    def _build_schema_analysis_prompt(
        self,
        sheet_name: str,
        sheet_data: List[List[str]],
        tables: List[Dict[str, Any]],
        context: Dict[str, Any],
    ) -> str:
        """Build comprehensive prompt for schema analysis."""

        prompt = f"""
Analyze this Excel sheet data and provide a comprehensive schema analysis.

SHEET NAME: {sheet_name}
FILE NAME: {context.get('file_name', 'Unknown')}

SAMPLE DATA (first 20 rows):
{self._format_data_for_prompt(sheet_data)}

DETECTED TABLES:
{json.dumps(tables, indent=2)}

CONTEXT INFORMATION:
- Related PDFs: {context.get('pdf_context', [])}
- Web source: {context.get('web_context', {}).get('source_url', 'None')}
- Directory files: {context.get('directory_context', [])}

Please provide a JSON response with this structure:
{{
    "sheet_type": "single_table|multiple_tables|metadata_sheet|mixed_content",
    "primary_purpose": "description of what this sheet contains",
    "tables": [
        {{
            "table_name": "descriptive name",
            "table_type": "data_table|lookup_table|summary_table|metadata",
            "columns": [
                {{
                    "name": "column_name",
                    "data_type": "text|number|date|boolean|mixed",
                    "description": "what this column represents",
                    "is_key": true/false,
                    "has_nulls": true/false,
                    "sample_values": ["val1", "val2", "val3"]
                }}
            ],
            "relationships": "description of how this table relates to others",
            "business_meaning": "what this table represents in business context"
        }}
    ],
    "analysis_queries": [
        "What questions can be answered from this data?",
        "What aggregations make sense?",
        "What filtering operations are useful?"
    ],
    "data_quality_notes": "observations about data quality, completeness, formatting",
    "recommended_processing": "how this data should be processed for analysis"
}}

Focus on understanding the business meaning and practical use cases for this data.
"""
        return prompt

    def _format_data_for_prompt(
        self, data: List[List[str]], max_width: int = 100
    ) -> str:
        """Format data for inclusion in prompt with strict token limits."""
        if not data:
            return "No data available"

        # Use only header + 2-3 sample rows to stay within context limits
        formatted_rows = []
        max_rows = min(4, len(data))  # Header + max 3 data rows

        for i, row in enumerate(data[:max_rows]):
            # Truncate long cells more aggressively
            truncated_row = [
                cell[:15] + "..." if len(cell) > 15 else cell
                for cell in row[:10]  # Max 10 columns
            ]
            row_label = "Header:" if i == 0 else f"Sample {i}:"
            formatted_rows.append(f"{row_label} {truncated_row}")

        return "\n".join(formatted_rows)

    def _create_fallback_schema(
        self, sheet_data: List[List[str]], tables: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """Create a basic fallback schema when LLM analysis fails."""

        if not sheet_data:
            return {"error": "No data to analyze"}

        # Simple heuristic-based schema
        headers = sheet_data[0] if sheet_data else []

        columns = []
        for i, header in enumerate(headers):
            if i < len(sheet_data[1:]) and sheet_data[1:]:
                sample_values = [
                    row[i] if i < len(row) else "" for row in sheet_data[1:]
                ]
                sample_values = [v for v in sample_values if v.strip()][:3]
            else:
                sample_values = []

            columns.append(
                {
                    "name": header or f"Column_{i+1}",
                    "data_type": "mixed",
                    "description": f"Data column {i+1}",
                    "sample_values": sample_values,
                }
            )

        return {
            "sheet_type": "single_table",
            "primary_purpose": "Data table (auto-detected)",
            "tables": [
                {
                    "table_name": "Main Table",
                    "table_type": "data_table",
                    "columns": columns,
                    "business_meaning": "Automatically detected data table",
                }
            ],
            "analysis_queries": ["Basic filtering and aggregation queries"],
            "fallback": True,
        }
