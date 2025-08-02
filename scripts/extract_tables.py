#!/usr/bin/env python3
import openpyxl
import pandas as pd
import os
import json
import yaml
from openai import OpenAI

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))


# ---------------------------
# 1. MERGED CELL HANDLING
# ---------------------------
def expand_merged_cells(ws):
    """Expand merged cells by filling them with the top-left value."""
    # Create a list of merged ranges to process
    merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in merged_ranges:
        # Get the top-left cell coordinate and value
        top_left_coord = merged_range.coord.split(":")[0]
        top_left_value = ws[top_left_coord].value

        # Unmerge the cells first
        ws.unmerge_cells(merged_range.coord)

        # Now fill all cells in the range with the top-left value
        for row in ws[merged_range.coord]:
            for cell in row:
                cell.value = top_left_value


def read_sheet_as_dataframe(file_path, sheet_name):
    """Read Excel sheet as DataFrame while expanding merged cells."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    expand_merged_cells(ws)

    max_row, max_col = ws.max_row, ws.max_column
    grid = [
        [ws.cell(row=r, column=c).value or "" for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]
    return pd.DataFrame(grid)


# ---------------------------
# 2. LLM HEADER INFERENCE
# ---------------------------
def infer_headers_multirow_llm(preview_rows):
    """
    Given the first few rows (list of lists), infer a flat list of headers.
    """
    raw_preview = "\n".join([", ".join(map(str, row)) for row in preview_rows])
    prompt = f"""
    The following rows represent the top of a table. The headers may span multiple rows:
    {raw_preview}

    Combine these rows into a single list of column names (flattened).
    Return ONLY a JSON array of column names, no explanation, no markdown formatting.
    Example: ["Column1", "Column2", "Column3"]
    """
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": "You are an expert at parsing table headers. Always return pure JSON arrays without any markdown formatting or code blocks.",
            },
            {"role": "user", "content": prompt},
        ],
        temperature=0,
    )
    content = response.choices[0].message.content.strip()
    # Remove any markdown formatting if present
    if content.startswith("```") and content.endswith("```"):
        lines = content.split("\n")
        content = "\n".join(lines[1:-1])
    return json.loads(content)


def generate_schema_with_llm(headers):
    """
    Generate a YAML schema describing each column's meaning using LLM.
    """
    prompt = f"""
    Given the following column headers:
    {headers}

    Describe each column's meaning in simple terms, return as YAML mapping:
    column_name: description

    Return ONLY the YAML content without any markdown formatting or code blocks.
    """
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "system",
                "content": "You are a schema documentation expert. Always return pure YAML without any markdown formatting or code blocks.",
            },
            {"role": "user", "content": prompt},
        ],
        temperature=0,
    )
    content = response.choices[0].message.content.strip()
    # Remove any markdown formatting if present
    if content.startswith("```") and content.endswith("```"):
        lines = content.split("\n")
        content = "\n".join(lines[1:-1])
    return yaml.safe_load(content)


# ---------------------------
# 3. TABLE EXTRACTION
# ---------------------------
def extract_tables_from_sheet(file_path, sheet_name, preview_rows=3):
    """
    Extract multiple tables from a sheet by splitting on empty rows.
    """
    df = read_sheet_as_dataframe(file_path, sheet_name)
    df = df.fillna("")

    # Identify empty rows to split tables
    split_points = df.index[
        df.apply(lambda row: all(v == "" for v in row), axis=1)
    ].tolist()

    tables = []
    start = 0
    for idx in split_points + [len(df)]:
        chunk = df.iloc[start:idx].dropna(how="all").reset_index(drop=True)
        if not chunk.empty:
            header_preview = chunk.head(preview_rows).values.tolist()
            try:
                headers = infer_headers_multirow_llm(header_preview)
            except Exception as e:
                print(f"[WARN] Header inference failed: {e}. Using default headers.")
                headers = [f"Column_{i}" for i in range(chunk.shape[1])]

            chunk = chunk.iloc[preview_rows:].reset_index(drop=True)

            # Handle header count mismatch
            if len(headers) != chunk.shape[1]:
                print(
                    f"[WARN] Header count mismatch: {len(headers)} headers vs {chunk.shape[1]} columns. Adjusting."
                )
                if len(headers) < chunk.shape[1]:
                    # Add missing headers
                    headers.extend(
                        [f"Column_{i}" for i in range(len(headers), chunk.shape[1])]
                    )
                else:
                    # Trim excess headers
                    headers = headers[: chunk.shape[1]]

            chunk.columns = headers
            tables.append(chunk)
        start = idx + 1

    return tables


def save_tables(tables, base_name, save_schema=False):
    for i, table in enumerate(tables, start=1):
        csv_path = f"{base_name}_table_{i}.csv"
        table.to_csv(csv_path, index=False)
        print(f"[INFO] Saved Table {i}: {csv_path}")

        if save_schema:
            schema = generate_schema_with_llm(list(table.columns))
            schema_path = f"{base_name}_table_{i}_schema.yaml"
            with open(schema_path, "w") as f:
                yaml.dump(schema, f)
            print(f"[INFO] Saved schema: {schema_path}")


def process_excel(file_path, save_schema=False):
    xls = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in xls.sheetnames:
        print(f"[INFO] Processing sheet: {sheet_name}")
        tables = extract_tables_from_sheet(file_path, sheet_name)
        base_name = f"{os.path.splitext(file_path)[0]}_{sheet_name}"
        save_tables(tables, base_name, save_schema)


# ---------------------------
# 4. ENTRY POINT
# ---------------------------
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Robust Multi-Table Extractor with LLM support."
    )
    parser.add_argument("file", help="Path to the Excel file.")
    parser.add_argument(
        "--schema", action="store_true", help="Generate YAML schema for columns."
    )
    args = parser.parse_args()

    process_excel(args.file, save_schema=args.schema)
